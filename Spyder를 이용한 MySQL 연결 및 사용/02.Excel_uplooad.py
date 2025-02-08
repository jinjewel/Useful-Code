import openpyxl

wb = openpyxl.load_workbook('C:/Users/User/Desktop/test.xlsx')

print(wb.sheetnames) # 시트 이름 출력

sheet_1 = wb['Sheet1'] # 시트 가지고 오기

print(sheet_1['A1'].value) # A1셀의 값을 가져오기
print(sheet_1['B1'].value) # B1셀의 값을 가져오기
print(sheet_1['E1'].value) # E1셀의 값을 가져오기, 없으면 None 반환
print()

# column (세로) 값들 읽기
for cols in sheet_1.iter_cols(): # 차례대로 A,B,C ... 증가
    for cell in cols: # 차례대로 1, 2, 3, ... 증가
        # A1, A2, ... , 마지막 행 데이터뽑고
        print(cell.value, end=" ")
    print()

print()        
# row (가로) 값들 읽기
for rows in sheet_1.iter_rows(): # 차례대로 A,B,C ... 증가
    for cell in rows: # 차례대로 1, 2, 3, ... 증가
        # A1, A2, ... , 마지막 행 데이터뽑고
        print(cell.value, end=" ")
    print()



